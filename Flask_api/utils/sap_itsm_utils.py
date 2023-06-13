import os
import time
#import pyrfc
import json
from datetime import datetime
from pymongo import MongoClient
import urllib.parse
import pandas as pd
import requests
import openpyxl
from ruamel.yaml import YAML

class Utilities():
    # A class to host the functional utility modules.

    def write_to_a_txt(self, write_data, file_name):

        # A function to write data to any sort of file.
        def_filename = open(file_name, 'w+')
        def_filename.write(write_data)
        def_filename.close()
        return
    
    def read_json_file(self, file_name):

        # A function to read values of a json file.
        d_filename = file_name
        d_filedata = open(d_filename, 'r')
        json_load = json.loads(d_filedata.read())
        return json_load
    
    def write_json_file(self, filename, data_to_write):
        
        # A function to write values to a json file
        # Serializing json
        json_object = json.dumps(data_to_write, indent=4)
 
        # Writing to sample.json
        with open(filename, "w") as outfile:
            outfile.write(json_object)
        
        return
    
    def convert_wa_to_list(self, wa_data):
        
        # A function to convert rfc data into a nested list
        table_length = len(wa_data)
        i = 0
        po_list = []
        
        # break down the raw data into nested lists for data processing
        for p_o in range(table_length):

            data_to_add = wa_data[i]['WA']
            data_split = data_to_add.split("|")
            po_list.append(data_split)
            i = i + 1
        
        return po_list
    
    def current_date(self):
        # A function to return the current time
        
        # Get the current time in seconds since the epoch
        current_time = time.time()

        # Convert the current time to a struct_time object
        time_struct = time.localtime(current_time)

        # Extract the year, month, and day from the struct_time object
        year = time_struct.tm_year
        month = time_struct.tm_mon
        day = time_struct.tm_mday

        # Format the date string in the desired format
        date_str = f"{year:04}-{month:02}-{day:02}"

        # Print the date string
        return date_str
    
    def date_range(self, month_number):
        
        import datetime
        
        # A function to generate a range of dates in a given month
        
        month_number = int(month_number)
        
        # Set the start and end dates
        start_date = datetime.date(2023, month_number, 1)
        end_date = datetime.date(2023, (month_number + 1), 1)

        # Generate a list of dates between the start and end dates
        date_list = [start_date + datetime.timedelta(days=x) for x in range((end_date - start_date).days)]

        # Convert each date to the desired format (yyyymmdd)
        date_list_formatted = [date.strftime('%Y%m%d') for date in date_list]
        
        # return the date range
        return date_list_formatted
    
    
    def time_range(self, minutes_before):
        
        # A function to generate a range of time values for the past variable minutes.
        # Get the current time
        now = datetime.datetime.now()

        # Calculate the start time (30 minutes ago)
        start_time = now - datetime.timedelta(minutes=30)

        # Create a list to store the time values
        time_list = []

        # Loop through the time range and append the time values to the list
        while start_time <= now:
            time_str = start_time.strftime("%H%M%S")
            time_list.append(time_str)
            start_time += datetime.timedelta(seconds=1)

        # Return the time list
        return time_list
    
    def read_yaml(self, yaml_file):
        # A function to read any input yaml file
        
        # loading the yaml object
        yaml = YAML()
        
        # reading the input file and returning the parameters
        config_file_read = open(yaml_file).read()
        config_parameters = yaml.load(config_file_read)
        
        # return the yaml contents
        return config_parameters
    
    def write_yaml(self, yaml_file, parameters):
        # A function to write any output to a yaml file
        
        # loading the yaml object
        yaml = YAML()
        
        # writing the yaml file
        file_write = open(yaml_file, "w")
        yaml.dump(parameters, file_write)
        file_write.close()
        
        return
    
    def send_requests(self, outputData, filename):
        
        print("---REQUESTS---START---")
        # A function to send json data using the requests library        
        sheaders = {'Content-Type': 'application/json'}
        
        # read from the EKS creds file
        filePath  ="/home/SAPITSM/finalsapdemo/config/eks_cred.json"
        
        serverData = self.read_json_file(filePath)
        
        sProtocol = str(serverData['serverProtocol'])
        sIP = str(serverData['serverIP'])
        sPort = str(serverData['serverPort'])
        sApi = str(serverData['serverEndpoint'])
        
        serverUrl = sProtocol + "://" + sIP + ":" + sPort + sApi
        
        print(serverUrl)
        
        try:
            print("-------------Try block ----------------")
            with open(filename, 'rb') as file:
                exception_file = {'file': file}
                output_data = {'json_data': json.dumps([outputData])}
                response = requests.post(serverUrl, files=exception_file, data=output_data, timeout=8)
            
            if response.status_code == 200:
                print('File and JSON data successfully sent.')
            else:
                print('An error occurred while sending the file and JSON data.')
            # with open(file_path, 'rb') as file:
            #     files = {'file': file}
                # print("---TRY---")
                # re = requests.post(serverUrl, json = outputData, files=files, headers = sheaders, verify=True, timeout=8)
            
            # payload = {'data': json.dumps(outputData)}
            # attachment = {'file': open(file_path, 'rb')}
            # re = requests.post(serverUrl, data=payload, files=attachment)
            
            # with open(file_path, 'rb') as file:
            #     file_data = file.read()
            
            # files = {
            #     'file': (file_path, file_data, 'text/plain')
            # }
            
            # payload = {
            #     'data': json.dumps(outputData)
            #     # 'files': files
            # }
            
            # print("---TRY---")
            # re = requests.post(serverUrl, json = outputData,  headers = sheaders, verify=True, timeout=8)
            
            # # re = requests.post(serverUrl, data=payload)
            # print(re)            
            
            # print("---SENT---")               
            # print(f"Status Code: {re.status_code}, Response: {re.json()}") 
            # re.raise_for_status()
        
        except requests.exceptions.RequestException as e:
            print("-----HTTP Exception-----")
            print('Failed to send the file:', e)
        except IOError as e:
            print("-----File Exception-----")
            print('File error:', e)
        except Exception as e:
            print("---Other EXCEPTIONS---")
            print("Error occured while sending data to the server.")
            print(e)
    
    def send_file_and_json(filename, json_data, url):
        # Got it from Jeevesh
        
        with open(filename, 'rb') as file:
            files = {'file': file}
            data = {'json_data': json.dumps([json_data])}
            response = requests.post(url, files=files, data=data)
        
        if response.status_code == 200:
            print('File and JSON data successfully sent.')
        else:
            print('An error occurred while sending the file and JSON data.')   
    
    def create_mysql_excel(self, file_name):
        # A function to create a mysql replacement excel temporarily

        # create a new workbook object
        wb = openpyxl.Workbook()

        # select the active sheet
        ws = wb.active

        # write some data to the sheet
        ws['A1'] = 'Purchase No'
        ws['B1'] = 'Alert time'        

        # save the workbook to a file
        wb.save(file_name)  
        wb.close()    
        
        return
    
    def read_excel_sheet(self, worksheet_name, sheet_name, column_name):
        wb = openpyxl.load_workbook(worksheet_name)  # Work Book
        ws = wb[sheet_name]  # Work Sheet
        column = ws[column_name]  # Column
        column_list = [column[x].value for x in range(len(column))]
        column_list.pop(0)
        
        wb.close
        
        return column_list 
    
    def get_rfc_connection(self, creds):

        # A function to do the rfc connection call to get data.
        rfc_conn = pyrfc.Connection(
            user = creds['user'],
            passwd = creds['passwd'],
            ashost = creds['ashost'],
            sysnr = creds['sysnr'],
            client = creds['client'],
        )
        return rfc_conn
    
    def get_rfc_read_table_for_fields(self, creds, query_table, fields, options):

        rfc_conn = self.get_rfc_connection(creds)

        rfc_data = rfc_conn.call('RFC_READ_TABLE', 
                   QUERY_TABLE=query_table, 
                   DELIMITER='|', 
                   FIELDS=fields,
                   OPTIONS=options)
        rfc_conn.close()

        return rfc_data
    
    def check_db(self):
        # A function to check if the database exists. 
        # If not, then create it.
        # A temporary substitute for DB
        parser = Utilities()
        excel_name = 'datastore/DB_temp.xlsx'

        if os.path.isfile(excel_name):
            print("File exists")
        else:
            print("File does not exist. Creating...")
            parser.create_mysql_excel(excel_name)
        
        return
        
